import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "BVA Test Cases"

headers = [
    "Test Case ID", 
    "Test Case Description", 
    "Test Case Procedure", 
    "Expected Results", 
    "Pre-conditions", 
    "Actual Result", 
    "Test Date", 
    "Tester", 
    "Note"
]

# Define styles
header_fill = PatternFill(start_color="7b8c38", end_color="7b8c38", fill_type="solid")
group1_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
group_font = Font(bold=True, color="000000")
alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'), 
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# Column widths
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 25
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 15
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15

# Define API structure
apis_service = [
    ("1.1. GET /api/reservations", [], "1"),
    ("1.2. GET /api/reservations/{id}", ["id"], "5"),
    ("1.3. POST /api/reservations", ["vehicleId", "userId", "startDatetime", "endDatetime", "purpose"], "21"),
    ("1.4. PUT /api/reservations/{id}", ["id", "startDatetime", "endDatetime", "purpose", "status"], "21"),
    ("1.5. PUT /api/reservations/{id}/status", ["id", "status"], "9"),
    ("1.6. DELETE /api/reservations/{id}", ["id"], "5"),
    ("1.7. GET /api/vehicles/{vehicleId}/reservations", ["vehicleId"], "5"),
    ("1.8. GET /api/vehicles/{vehicleId}/group", ["vehicleId"], "5"),
    ("1.9. GET /api/users/{userId}/vehicles", ["userId"], "5"),
    ("1.10. GET /api/users/{userId}/vehicles/{vehicleId}/reservations", ["userId", "vehicleId"], "9"),
    ("1.11. GET /api/availability", ["vehicleId", "start", "end"], "13"),
    ("1.12. GET /api/reservations/{id}/checkpoints", ["id"], "5"),
    ("1.13. POST /api/reservations/{id}/checkpoints", ["id", "type", "issuedBy"], "13"),
    ("1.14. POST /api/reservations/checkpoints/scan", ["token", "lat", "lng"], "13"),
    ("1.15. POST /api/reservations/checkpoints/{id}/sign", ["id", "signer"], "9"),
    ("1.16. GET /api/fairness/vehicles/{vehicleId}", ["vehicleId", "rangeDays"], "9"),
    ("1.17. POST /api/fairness/vehicles/{vehicleId}/suggest", ["vehicleId", "userId"], "9")
]

apis_admin = [
    ("2.1. POST /api/admin/auth/login", ["username", "password"], "9"),
    ("2.2. GET /api/admin/reservations", [], "1"),
    ("2.3. GET /api/admin/reservations/{id}", ["id"], "5"),
    ("2.4. PUT /api/admin/reservations/{id}", ["id", "status", "startDatetime", "endDatetime", "vehicleId", "userId"], "25"),
    ("2.5. DELETE /api/admin/reservations/{id}", ["id"], "5"),
    ("2.6. POST /api/admin/reservations/sync", ["reservationId", "status", "startDatetime", "endDatetime", "vehicleId", "userId"], "25"),
    ("2.7. GET /api/admin/reservations/manage", [], "1"),
    ("2.8. POST /api/admin/reservations/manage", ["vehicleId", "userId", "startDatetime", "endDatetime", "status"], "21"),
    ("2.9. PUT /api/admin/reservations/manage/{id}", ["id", "startDatetime", "endDatetime", "status"], "17"),
    ("2.10. DELETE /api/admin/reservations/manage/{id}", ["id"], "5"),
    ("2.11. GET /api/admin/vehicles", [], "1"),
    ("2.12. GET /api/reservations (Proxy)", [], "1"),
    ("2.13. GET /api/reservations/{id} (Proxy)", ["id"], "5"),
    ("2.14. POST /api/reservations (Proxy)", ["vehicleId", "userId", "startDatetime", "endDatetime", "purpose"], "21"),
    ("2.15. PUT /api/reservations/{id} (Proxy)", ["id", "startDatetime", "endDatetime", "purpose", "status"], "21"),
    ("2.16. PUT /api/reservations/{id}/status (Proxy)", ["id", "status"], "9"),
    ("2.17. DELETE /api/reservations/{id} (Proxy)", ["id"], "5"),
    ("2.18. GET /api/test/user/{userId}", ["userId"], "5")
]

def add_group_header(text):
    ws.append(["", text, "---", "---", "---", "", "", "", ""])
    row = ws[ws.max_row]
    for cell in row:
        cell.fill = group1_fill
        cell.font = group_font
        cell.alignment = alignment
        cell.border = thin_border

def write_bva_rows(prefix, start_idx, api_name, params, total_tc, precondition="Hệ thống sẵn sàng"):
    # Add API header row
    n = len(params)
    api_header = f"▶ {api_name} (n={n} -> {total_tc} TC)"
    nom_str = "Nominal: " + ", ".join([f"{p}=nom" for p in params]) if params else "Nominal: n/a"
    
    ws.append(["", api_header, nom_str, "---", "---", "", "", "", ""])
    row = ws[ws.max_row]
    for cell in row:
        cell.fill = group1_fill
        cell.font = group_font
        cell.border = thin_border
    
    tc_count = 1
    
    # helper for formatting row
    def add_tc(desc, proc, expected):
        nonlocal tc_count
        tc_id = f"{prefix}_{start_idx:02d}_{tc_count:03d}"
        ws.append([tc_id, desc, proc, expected, precondition, "", "", "", ""])
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            cell.alignment = alignment
        tc_count += 1

    # TC 1: All = nom
    add_tc("BVA: All = nom", 
           "Truyền tất cả " + ("tham số bằng giá trị nominal (nom)" if params else "không tham số"), 
           "200 OK")
    
    # Bounds for parameters
    for p in params:
        add_tc(f"BVA: {p} = min", f"Truyền {p} = min (giá trị nhỏ nhất hợp lệ)", "200 OK / 400")
        add_tc(f"BVA: {p} = min+1", f"Truyền {p} = min+1", "200 OK")
        add_tc(f"BVA: {p} = max-1", f"Truyền {p} = max-1", "200 OK")
        add_tc(f"BVA: {p} = max", f"Truyền {p} = max (giá trị chuẩn lớn nhất)", "200 OK / 400")

add_group_header("▶ NHÓM 1: RESERVATION SERVICE - 17 APIs")
for i, api in enumerate(apis_service):
    write_bva_rows("TC_RES", i+1, api[0], api[1], api[2])

add_group_header("▶ NHÓM 2: RESERVATION ADMIN SERVICE - 18 APIs")
for i, api in enumerate(apis_admin):
    write_bva_rows("TC_ADM", i+1, api[0], api[1], api[2])

wb.save("ECOCSS_BVA_TestCases_4n_plus_1.xlsx")
print("Tạo file ECOCSS_BVA_TestCases_4n_plus_1.xlsx thành công!")
