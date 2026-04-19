# ECOCSS Function 5 - Manual Test Case Generator
# Creates a multi-sheet Excel workbook matching the Group Management template.
# Run: python generate_ecocss_manual_test.py

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import os

# ──────────────────────────────────────────────
# 1.  STYLE HELPERS
# ──────────────────────────────────────────────

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)


def apply_header_style(cell, bg_hex="1F4E79", font_hex="FFFFFF",
                        bold=True, size=11, align="center"):
    cell.font = Font(bold=bold, color=font_hex, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=True)
    cell.border = make_border("thin")


def apply_subheader_style(cell, bg_hex="2E75B6", font_hex="FFFFFF",
                           bold=True, size=10):
    cell.font = Font(bold=bold, color=font_hex, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
    cell.border = make_border("thin")


def apply_data_style(cell, align="left", wrap=True, bold=False,
                      size=10):
    cell.font = Font(bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=wrap)
    cell.border = make_border("thin")


def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width


def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height


# ──────────────────────────────────────────────
# 2.  DATA  (from ECOCSS_Test_Case_Registry_Function5.tsv)
# ──────────────────────────────────────────────

RESERVATION_SERVICE_TCS = [
    # TC_ID, Function Name, Requirement Name, Method, Path, Port, Precondition, Test Steps, Expected Result, Priority
    ("ECOCSS-F5-001", "getAllReservations",    "Lấy toàn bộ đặt chỗ (booking chính)",     "GET",  "/api/reservations",                              8086, "Service 8086 chạy; MySQL booking có dữ liệu",                          "Mở Postman GET baseUrl8086/api/reservations; Send",                                                      "200; JSON array; mỗi phần tử có reservationId vehicleId status",                             "High"),
    ("ECOCSS-F5-002", "getReservationById",   "Lấy chi tiết đặt chỗ theo ID",              "GET",  "/api/reservations/{id}",                         8086, "Có reservationId hợp lệ",                                              "GET với {{reservationId}}",                                                                              "200; object khớp id URL",                                                                             "High"),
    ("ECOCSS-F5-003", "createReservation",     "Tạo đặt chỗ mới",                           "POST", "/api/reservations",                               8086, "Có vehicleId userId hợp lệ",                                            "POST body JSON theo collection",                                                                        "200 hoặc 201; có reservationId; status BOOKED",                                                       "High"),
    ("ECOCSS-F5-004", "updateReservation",     "Cập nhật đặt chỗ",                           "PUT",  "/api/reservations/{id}",                         8086, "Có reservation tồn tại",                                                "PUT body theo collection",                                                                              "200; field cập nhật khớp body",                                                                      "High"),
    ("ECOCSS-F5-005", "updateReservationStatus","Cập nhật trạng thái đặt chỗ",             "PUT",  "/api/reservations/{id}/status",                  8086, "Có reservation; status hợp lệ",                                        "PUT ?status=IN_USE",                                                                                    "200; status response khớp query",                                                                   "High"),
    ("ECOCSS-F5-006", "deleteReservation",     "Xóa đặt chỗ",                               "DELETE","/api/reservations/{id}",                      8086, "Có id cần xóa",                                                         "DELETE",                                                                                                "200 hoặc 204",                                                                                       "High"),
    ("ECOCSS-F5-007", "getVehicleReservations","Lịch đặt theo xe",                           "GET",  "/api/vehicles/{vehicleId}/reservations",          8086, "Có vehicleId",                                                          "GET",                                                                                                   "200; array; vehicleId khớp",                                                                        "High"),
    ("ECOCSS-F5-008", "getVehicleGroupInfo",   "Thông tin nhóm sở hữu theo xe",             "GET",  "/api/vehicles/{vehicleId}/group",                8086, "Group service sẵn sàng",                                                 "GET; bật Authorization nếu cần",                                                                        "200; có groupId members",                                                                            "High"),
    ("ECOCSS-F5-009", "getUserVehicles",        "Danh sách xe của người dùng",               "GET",  "/api/users/{userId}/vehicles",                   8086, "User có group",                                                         "GET; JWT nếu cần",                                                                                     "200; array",                                                                                        "High"),
    ("ECOCSS-F5-010", "getUserVehicleReservations","Lịch đặt theo user và xe",            "GET",  "/api/users/{userId}/vehicles/{vehicleId}/reservations", 8086, "Có dữ liệu",    "GET",                                                                                                   "200; array",                                                                                        "High"),
    ("ECOCSS-F5-011", "checkAvailability",     "Kiểm tra khả dụng slot",                    "GET",  "/api/availability",                               8086, "Query vehicleId start end",                                            "GET với query",                                                                                        "200; boolean",                                                                                      "High"),
    ("ECOCSS-F5-012", "listReservationCheckpoints","Liệt kê checkpoint",                 "GET",  "/api/reservations/{id}/checkpoints",              8086, "Có reservation",                                                        "GET",                                                                                                   "200; array",                                                                                        "High"),
    ("ECOCSS-F5-013", "issueReservationCheckpoint","Phát hành checkpoint",                 "POST", "/api/reservations/{id}/checkpoints",              8086, "Có reservation",                                                        "POST body type issuedBy",                                                                               "200 hoặc 201",                                                                                      "High"),
    ("ECOCSS-F5-014", "scanCheckpoint",         "Quét checkpoint",                            "POST", "/api/reservations/checkpoints/scan",              8086, "Có token hợp lệ",                                                       "POST token lat lng",                                                                                    "200 hoặc 201",                                                                                      "High"),
    ("ECOCSS-F5-015", "signCheckpoint",         "Ký checkpoint",                              "POST", "/api/reservations/checkpoints/{id}/sign",        8086, "Có checkpointId",                                                        "POST body signer",                                                                                      "200 hoặc 201",                                                                                      "High"),
    ("ECOCSS-F5-016", "getFairnessSummary",     "Tóm tắt công bằng",                          "GET",  "/api/fairness/vehicles/{vehicleId}",             8086, "Có vehicleId",                                                          "GET ?rangeDays=30",                                                                                    "200; object",                                                                                       "High"),
    ("ECOCSS-F5-017", "suggestFairnessSlot",   "Gợi ý slot công bằng",                      "POST", "/api/fairness/vehicles/{vehicleId}/suggest",     8086, "Có vehicleId userId",                                                     "POST body JSON",                                                                                       "200 hoặc 201",                                                                                      "High"),
]

ADMIN_SERVICE_TCS = [
    ("ECOCSS-F5-018", "adminLogin",                   "Đăng nhập admin",                     "POST",   "/api/admin/auth/login",                   8087, "Có tài khoản admin",                       "POST username password",                               "200 hoặc 201; có token",                            "High"),
    ("ECOCSS-F5-019", "getAllAdminReservations",       "Danh sách reservation admin",          "GET",    "/api/admin/reservations",                  8087, "Service 8087 chạy",                          "GET",                                                "200; array DTO admin",                            "High"),
    ("ECOCSS-F5-020", "getAdminReservationById",       "Chi tiết reservation admin",           "GET",    "/api/admin/reservations/{id}",             8087, "Có id",                                     "GET",                                                 "200 hoặc 404",                                    "High"),
    ("ECOCSS-F5-021", "updateAdminReservation",        "Cập nhật reservation admin",           "PUT",    "/api/admin/reservations/{id}",             8087, "Có id",                                     "PUT body",                                            "200",                                             "High"),
    ("ECOCSS-F5-022", "deleteAdminReservation",        "Xóa reservation admin",                 "DELETE", "/api/admin/reservations/{id}",             8087, "Có id",                                     "DELETE",                                              "200 hoặc 204",                                    "High"),
    ("ECOCSS-F5-023", "syncReservationFromBooking",    "Đồng bộ booking",                       "POST",   "/api/admin/reservations/sync",             8087, "Payload đầy đủ",                             "POST JSON",                                            "200 hoặc 201",                                    "High"),
    ("ECOCSS-F5-024", "getManageReservations",         "Danh sách manage",                      "GET",    "/api/admin/reservations/manage",            8087, "—",                                         "GET",                                                 "200; array manage view",                          "High"),
    ("ECOCSS-F5-025", "createManageReservation",       "Tạo lịch manage",                       "POST",   "/api/admin/reservations/manage",            8087, "vehicleId userId hợp lệ",                   "POST body",                                            "200 hoặc 201; message id",                       "High"),
    ("ECOCSS-F5-026", "updateManageReservation",       "Cập nhật lịch manage",                  "PUT",    "/api/admin/reservations/manage/{id}",        8087, "Có id",                                     "PUT body",                                             "200; message",                                    "High"),
    ("ECOCSS-F5-027", "deleteManageReservation",       "Xóa lịch manage",                       "DELETE", "/api/admin/reservations/manage/{id}",        8087, "Có id",                                     "DELETE",                                               "200 hoặc 204",                                    "High"),
    ("ECOCSS-F5-028", "getAdminVehicles",               "Danh sách xe admin",                    "GET",    "/api/admin/vehicles",                        8087, "—",                                         "GET",                                                 "200; array",                                      "High"),
    ("ECOCSS-F5-029", "proxyGetAllReservations",       "Proxy GET tất cả",                      "GET",    "/api/reservations",                          8087, "8087 chạy",                                 "GET baseUrl8087",                                      "200; array",                                      "High"),
    ("ECOCSS-F5-030", "proxyGetReservationById",        "Proxy GET theo id",                     "GET",    "/api/reservations/{id}",                    8087, "Có id",                                     "GET",                                                 "200 hoặc 404",                                    "High"),
    ("ECOCSS-F5-031", "proxyCreateReservation",         "Proxy POST tạo",                        "POST",   "/api/reservations",                          8087, "Body hợp lệ",                               "POST",                                                "200 hoặc 201",                                    "High"),
    ("ECOCSS-F5-032", "proxyUpdateReservation",         "Proxy PUT",                              "PUT",    "/api/reservations/{id}",                    8087, "Có id",                                     "PUT",                                                 "200",                                             "High"),
    ("ECOCSS-F5-033", "proxyUpdateReservationStatus",   "Proxy PUT status",                      "PUT",    "/api/reservations/{id}/status",              8087, "Có id; status query",                       "PUT",                                                 "200",                                             "High"),
    ("ECOCSS-F5-034", "proxyDeleteReservation",         "Proxy DELETE",                           "DELETE", "/api/reservations/{id}",                    8087, "Có id",                                     "DELETE",                                              "200 hoặc 204",                                    "High"),
    ("ECOCSS-F5-035", "getTestUserById",                "Test user debug",                        "GET",    "/api/test/user/{userId}",                   8087, "Có userId",                                 "GET",                                                 "200; object",                                     "Medium"),
]

# ──────────────────────────────────────────────
# 3.  SHEET 0 – GUIDELINE
# ──────────────────────────────────────────────

def create_guideline_sheet(wb):
    ws = wb.active
    ws.title = "Guideline"

    # Title row
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "ECOCSS - MANUAL TEST CASE GUIDELINES"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtitle
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = "EV Co-ownership Cost Sharing System — Function 5: Reservation & Admin"
    sub.font = Font(italic=True, size=11, color="FFFFFF", name="Calibri")
    sub.fill = PatternFill("solid", fgColor="2E75B6")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Header row
    ws.row_dimensions[3].height = 25
    headers = ["#", "Item", "Description / Example"]
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.value = hdr
        apply_header_style(c, bg_hex="BDD7EE")

    # Data rows
    guide_items = [
        (1, "Purpose",         "This workbook defines manual test cases for ReservationService (port 8086) and ReservationAdminService (port 8087). Use it together with the BVA detail sheet."),
        (2, "Prerequisites",   "- Postman or any REST client\n- Service running on correct port\n- MySQL DB seeded with test data\n- Admin credentials if testing auth endpoints"),
        (3, "Test Execution",  "1. Open the appropriate Service sheet.\n2. Follow Test Steps exactly.\n3. Record Actual Result.\n4. Set Status: PASSED / FAILED / PENDING / N/A."),
        (4, "Status Codes",    "- PASSED  = Expected result matches actual\n- FAILED  = Actual differs from expected\n- PENDING = Not yet executed\n- N/A     = Not applicable"),
        (5, "Priority",        "- High   = P0 - Must pass before release\n- Medium = P1 - Should pass\n- Low    = P2 - Nice to have"),
        (6, "Round Tracking", "Use the Statistics sheet to track Round 1 / Round 2 / Round 3 results."),
        (7, "Reporting",      "After each test round, update Statistics sheet. Escalate FAILED items immediately."),
    ]

    row = 4
    for item in guide_items:
        ws.row_dimensions[row].height = 50
        ws.cell(row=row, column=1).value = item[0]
        ws.cell(row=row, column=2).value = item[1]
        ws.cell(row=row, column=3).value = item[2]

        for col in range(1, 4):
            c = ws.cell(row=row, column=col)
            c.font = Font(size=10, name="Calibri")
            c.border = make_border()
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if col == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font = Font(bold=True, size=10, name="Calibri")
                c.fill = PatternFill("solid", fgColor="DEEAF1")
            if col == 2:
                c.fill = PatternFill("solid", fgColor="EBF3FB")
                c.font = Font(bold=True, size=10, name="Calibri")
        row += 1

    # Col widths
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 80

    # Freeze header
    ws.freeze_panes = "A4"


# ──────────────────────────────────────────────
# 4.  SHEET 1 – COVER
# ──────────────────────────────────────────────

def create_cover_sheet(wb):
    ws = wb.create_sheet("Cover")

    # Row 1 – banner
    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 40
    c = ws["A1"]
    c.value = "ECOCSS — FUNCTION 5: RESERVATION & ADMIN  |  MANUAL TEST CASE WORKBOOK"
    c.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 – subtitle
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 22
    c = ws["A2"]
    c.value = "EV Co-ownership Cost Sharing System"
    c.font = Font(italic=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])  # row 3 blank

    # Info block header
    ws.row_dimensions[4].height = 22
    ws.merge_cells("A4:H4")
    c = ws["A4"]
    c.value = "[i]  PROJECT INFORMATION"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")

    info_rows = [
        ("Feature",         "Function 5 - Reservation & Admin Management"),
        ("Services",        "ReservationService (8086) | ReservationAdminService (8087)"),
        ("Test Type",       "Manual API Testing"),
        ("Test Technique",  "Equivalence Partitioning + Boundary Value Analysis (BVA)"),
        ("Number of TCs",    f"{len(RESERVATION_SERVICE_TCS) + len(ADMIN_SERVICE_TCS)} test cases"),
        ("ReservationService TCs",  f"{len(RESERVATION_SERVICE_TCS)} cases"),
        ("AdminService TCs",       f"{len(ADMIN_SERVICE_TCS)} cases"),
        ("Created",         "2026-04-04"),
        ("Status",          "DRAFT"),
    ]

    for idx, (label, value) in enumerate(info_rows, start=5):
        ws.row_dimensions[idx].height = 20
        ws.merge_cells(f"A{idx}:B{idx}")
        ws.merge_cells(f"C{idx}:H{idx}")
        lc = ws.cell(row=idx, column=1)
        lc.value = label
        lc.font = Font(bold=True, size=10, name="Calibri")
        lc.fill = PatternFill("solid", fgColor="DEEAF1")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = make_border()

        vc = ws.cell(row=idx, column=3)
        vc.value = value
        vc.font = Font(size=10, name="Calibri")
        vc.fill = PatternFill("solid", fgColor="FFFFFF")
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = make_border()

    # Round summary header
    round_start = 5 + len(info_rows) + 1
    ws.row_dimensions[round_start].height = 22
    ws.merge_cells(f"A{round_start}:H{round_start}")
    c = ws.cell(row=round_start, column=1)
    c.value = "[i]  TEST ROUND SUMMARY"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")

    # Round table headers
    rh = round_start + 1
    ws.row_dimensions[rh].height = 22
    round_headers = ["Round", "Passed", "Failed", "Pending", "N/A", "Executed By", "Date", "Notes"]
    for col, h in enumerate(round_headers, 1):
        c = ws.cell(row=rh, column=col)
        c.value = h
        apply_header_style(c, bg_hex="2E75B6")

    # Round data rows
    for i in range(1, 4):
        rr = rh + i
        ws.row_dimensions[rr].height = 20
        ws.cell(row=rr, column=1).value = f"Round {i}"
        ws.cell(row=rr, column=6).value = ""
        ws.cell(row=rr, column=7).value = ""
        ws.cell(row=rr, column=8).value = ""
        for col in range(1, 9):
            c = ws.cell(row=rr, column=col)
            c.font = Font(size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor="FFFFFF")
            c.border = make_border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            if col == 1:
                c.fill = PatternFill("solid", fgColor="DEEAF1")
                c.font = Font(bold=True, size=10, name="Calibri")

    # Service sheets navigation
    nav_start = rh + 4
    ws.row_dimensions[nav_start].height = 22
    ws.merge_cells(f"A{nav_start}:H{nav_start}")
    c = ws.cell(row=nav_start, column=1)
    c.value = "[i]  SHEETS IN THIS WORKBOOK"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")

    nav_items = [
        ("Guideline",       "How to use this workbook"),
        ("Cover",           "Project overview & round summary (this sheet)"),
        ("ReservationService", "17 test cases — port 8086"),
        ("ReservationAdminService", "18 test cases — port 8087"),
        ("Statistics",      "Summary of all rounds & pass/fail rates"),
    ]
    for idx, (sheet, desc) in enumerate(nav_items, start=nav_start + 1):
        ws.row_dimensions[idx].height = 20
        ws.merge_cells(f"A{idx}:B{idx}")
        ws.merge_cells(f"C{idx}:H{idx}")
        for col in range(1, 9):
            c = ws.cell(row=idx, column=col)
            c.font = Font(size=10, name="Calibri")
            c.fill = PatternFill("solid", fgColor="FFFFFF")
            c.border = make_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=idx, column=1).value = sheet
        ws.cell(row=idx, column=1).font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=idx, column=1).fill = PatternFill("solid", fgColor="DEEAF1")
        ws.cell(row=idx, column=3).value = desc

    # Col widths
    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["C"].width = 45


# ──────────────────────────────────────────────
# 5.  SHARED TEST-CASE TABLE BUILDER
# ──────────────────────────────────────────────

TC_HEADERS = [
    "TC_ID", "Service", "#", "Function Name", "Requirement Name",
    "Method", "Path", "Port", "Precondition", "Test Steps",
    "Expected Result", "Priority",
    "Status", "Actual Result", "Executed By", "Date", "Notes"
]

# Color map for Status
STATUS_COLORS = {
    "High":   "FF0000",   # Red   – must pass
    "Medium": "FFC000",   # Amber – should pass
    "Low":    "00B050",   # Green – nice to have
}


def create_service_sheet(wb, sheet_name, service_tcs):
    ws = wb.create_sheet(sheet_name)

    # ── Title banner ──────────────────────────────
    last_col = len(TC_HEADERS)
    ws.merge_cells(f"A1:{get_column_letter(last_col)}1")
    ws.row_dimensions[1].height = 32
    c = ws["A1"]
    c.value = f"ECOCSS – FUNCTION 5  |  {sheet_name.upper()}  |  MANUAL TEST CASES"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sub-banner: service info ─────────────────
    ws.merge_cells(f"A2:{get_column_letter(last_col)}2")
    ws.row_dimensions[2].height = 20
    svc_port = "8086" if sheet_name == "ReservationService" else "8087"
    svc_desc = ("Core reservation APIs — user-facing booking operations"
                if sheet_name == "ReservationService"
                else "Admin & proxy APIs — management dashboard")
    c = ws["A2"]
    c.value = f"Service: {sheet_name}  |  Port: {svc_port}  |  {svc_desc}"
    c.font = Font(italic=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Round tracker mini-bar (row 3) ───────────
    ws.merge_cells(f"A3:{get_column_letter(last_col)}3")
    ws.row_dimensions[3].height = 18
    c = ws["A3"]
    c.value = "Round tracker:   Round 1: __   Round 2: __   Round 3: __"
    c.font = Font(size=9, italic=True, name="Calibri", color="595959")
    c.alignment = Alignment(horizontal="right", vertical="center")

    # ── Header row ──────────────────────────────
    ws.row_dimensions[4].height = 35
    for col, hdr in enumerate(TC_HEADERS, 1):
        c = ws.cell(row=4, column=col)
        c.value = hdr
        apply_header_style(c, bg_hex="1F4E79")

    # ── Data rows ───────────────────────────────
    for row_idx, tc in enumerate(service_tcs, start=5):
        ws.row_dimensions[row_idx].height = 55
        tc_id, func, req, method, path, port, precond, steps, expected, priority = tc

        row_data = [
            tc_id,        # TC_ID
            sheet_name,   # Service
            row_idx - 4,  # #
            func,         # Function Name
            req,          # Requirement Name
            method,       # Method
            path,         # Path
            port,         # Port
            precond,      # Precondition
            steps,        # Test Steps
            expected,     # Expected Result
            priority,     # Priority
            "",           # Status (blank – to fill)
            "",           # Actual Result (blank)
            "",           # Executed By (blank)
            "",           # Date (blank)
            "",           # Notes (blank)
        ]

        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col)
            c.value = val
            c.border = make_border()
            c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Priority color coding
            if col == 12:  # Priority column
                prio_color = STATUS_COLORS.get(priority, "000000")
                c.font = Font(bold=True, size=9, color=prio_color, name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 1:  # TC_ID
                c.font = Font(bold=True, size=9, name="Calibri")
                c.fill = PatternFill("solid", fgColor="DEEAF1")
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 6:  # Method
                method_colors = {"GET": "70AD47", "POST": "4472C4", "PUT": "ED7D31", "DELETE": "FF0000"}
                c.font = Font(bold=True, size=9, color=method_colors.get(method, "000000"), name="Calibri")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = Font(size=9, name="Calibri")

    # ── Column widths ────────────────────────────
    col_widths = [14, 24, 5, 22, 30, 8, 40, 8, 35, 40, 40, 10, 10, 35, 16, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze at header ─────────────────────────
    ws.freeze_panes = "A5"

    # ── Auto-filter on header row ───────────────
    ws.auto_filter.ref = f"A4:{get_column_letter(last_col)}{4 + len(service_tcs)}"

    return ws


# ──────────────────────────────────────────────
# 6.  SHEET – STATISTICS
# ──────────────────────────────────────────────

def create_statistics_sheet(wb):
    ws = wb.create_sheet("Statistics")

    # Title
    ws.merge_cells("A1:L1")
    ws.row_dimensions[1].height = 32
    c = ws["A1"]
    c.value = "ECOCSS – FUNCTION 5  |  TEST EXECUTION STATISTICS"
    c.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 20
    c = ws["A2"]
    c.value = "ReservationService (17 TCs)  +  ReservationAdminService (18 TCs)  =  35 Total TCs"
    c.font = Font(italic=True, size=10, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="2E75B6")
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.append([])  # row 3 blank

    # ── Per-service summary ─────────────────────
    ws.row_dimensions[4].height = 22
    ws.merge_cells("A4:L4")
    c = ws["A4"]
    c.value = "[!]  PER-SERVICE SUMMARY"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="1F4E79")
    c.alignment = Alignment(horizontal="left", vertical="center")

    svc_headers = ["Service", "Total TCs", "High", "Medium", "Low",
                   "Round 1 Passed", "Round 1 Failed", "Round 1 Pending",
                   "Round 2 Passed", "Round 2 Failed", "Round 2 Pending",
                   "Notes"]
    ws.row_dimensions[5].height = 22
    for col, h in enumerate(svc_headers, 1):
        c = ws.cell(row=5, column=col)
        c.value = h
        apply_header_style(c, bg_hex="2E75B6")

    services_summary = [
        ("ReservationService",       17, 17, 0, 0, "", "", "", "", "", "", ""),
        ("ReservationAdminService",   18,  1, 0, 17, "", "", "", "", "", "", ""),
        ("TOTAL",                     35, 18, 0, 17, "", "", "", "", "", "", ""),
    ]
    for idx, row_data in enumerate(services_summary, start=6):
        ws.row_dimensions[idx].height = 22
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=idx, column=col)
            c.value = val
            c.border = make_border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=10, name="Calibri", bold=(col == 1 or idx == 8))
            if idx == 8:  # TOTAL row
                c.fill = PatternFill("solid", fgColor="DEEAF1")
                c.font = Font(bold=True, size=10, name="Calibri")

    ws.append([])  # row 9 blank

    # ── Priority breakdown ───────────────────────
    ws.row_dimensions[10].height = 22
    ws.merge_cells("A10:L10")
    c = ws["A10"]
    c.value = "[!] HIGH PRIORITY TCs -- Must pass before release"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="C00000")
    c.alignment = Alignment(horizontal="left", vertical="center")

    prio_headers = ["TC_ID", "Function Name", "Service", "Method", "Path",
                    "Priority", "Round 1", "R1 Result", "Round 2", "R2 Result", "Round 3", "R3 Result"]
    ws.row_dimensions[11].height = 22
    for col, h in enumerate(prio_headers, 1):
        c = ws.cell(row=11, column=col)
        c.value = h
        apply_header_style(c, bg_hex="E00000")

    # Collect all High priority TCs
    high_tcs = [(t[0], t[1], sheet_name, t[3], t[4])
                for sheet_name, tcs in [("ReservationService", RESERVATION_SERVICE_TCS),
                                         ("ReservationAdminService", ADMIN_SERVICE_TCS)]
                for t in tcs if t[9] == "High"]

    for idx, tc in enumerate(high_tcs, start=12):
        ws.row_dimensions[idx].height = 20
        row_vals = [tc[0], tc[1], tc[2], tc[3], tc[4], "High",
                    "Round 1", "", "Round 2", "", "Round 3", ""]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=idx, column=col)
            c.value = val
            c.border = make_border()
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center" if col > 5 else "left",
                                    vertical="center")
            if col <= 5:
                c.fill = PatternFill("solid", fgColor="FFE0E0")

    ws.append([])

    # ── Medium priority ─────────────────────────
    med_start = 12 + len(high_tcs) + 1
    ws.row_dimensions[med_start].height = 22
    ws.merge_cells(f"A{med_start}:L{med_start}")
    c = ws.cell(row=med_start, column=1)
    c.value = "[!] MEDIUM PRIORITY TCs -- Should pass"
    c.font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    c.fill = PatternFill("solid", fgColor="BF8F00")
    c.alignment = Alignment(horizontal="left", vertical="center")

    med_tcs = [(t[0], t[1], sheet_name, t[3], t[4])
               for sheet_name, tcs in [("ReservationService", RESERVATION_SERVICE_TCS),
                                        ("ReservationAdminService", ADMIN_SERVICE_TCS)]
               for t in tcs if t[9] == "Medium"]

    ws.row_dimensions[med_start + 1].height = 22
    for col, h in enumerate(prio_headers, 1):
        c = ws.cell(row=med_start + 1, column=col)
        c.value = h
        apply_header_style(c, bg_hex="FFC000")

    for idx, tc in enumerate(med_tcs, start=med_start + 2):
        ws.row_dimensions[idx].height = 20
        row_vals = [tc[0], tc[1], tc[2], tc[3], tc[4], "Medium",
                    "Round 1", "", "Round 2", "", "Round 3", ""]
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=idx, column=col)
            c.value = val
            c.border = make_border()
            c.font = Font(size=9, name="Calibri")
            c.alignment = Alignment(horizontal="center" if col > 5 else "left",
                                    vertical="center")
            if col <= 5:
                c.fill = PatternFill("solid", fgColor="FFF2CC")

    # Col widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 10
    for col in "GHIJKL":
        ws.column_dimensions[col].width = 14

    ws.freeze_panes = "A5"


# ──────────────────────────────────────────────
# 7.  MAIN
# ──────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()

    # 1. Guideline
    create_guideline_sheet(wb)

    # 2. Cover
    create_cover_sheet(wb)

    # 3. ReservationService sheet
    create_service_sheet(wb, "ReservationService", RESERVATION_SERVICE_TCS)

    # 4. ReservationAdminService sheet
    create_service_sheet(wb, "ReservationAdminService", ADMIN_SERVICE_TCS)

    # 5. Statistics
    create_statistics_sheet(wb)

    # Save
    output_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "ECOCSS_Function5_Manual_TestCase.xlsx"
    )
    wb.save(output_path)
    print("")
    print("[OK] File created:", output_path)
    print("     Sheets:", [s.title for s in wb.worksheets])
    print("     ReservationService TCs :", len(RESERVATION_SERVICE_TCS))
    print("     ReservationAdminService TCs:", len(ADMIN_SERVICE_TCS))
    print("     Total TCs              :", len(RESERVATION_SERVICE_TCS) + len(ADMIN_SERVICE_TCS))


if __name__ == "__main__":
    main()
